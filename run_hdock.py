"""
run_hdock.py  –  Birleşik Docking & Rapor Pipeline'ı
=====================================================
- Klasördeki .pdb dosyalarından adında "APTAMER" geçeni otomatik ligand olarak seçer.
- Geri kalan tüm .pdb dosyalarını receptor olarak alır.
- HDOCK ile docking yapar (WSL → fallback: native).
- Sonuçları parse eder ve profesyonel tek sayfalık bir Excel raporu üretir.

Kullanım:
  python run_hdock.py                       # Otomatik mod
  python run_hdock.py -l MyAptamer.pdb      # Ligandı elle belirt
  python run_hdock.py -r hedef1.pdb hedef2.pdb  # Sadece belirli receptorlar
  python run_hdock.py --top 10              # En iyi 10 sonucu al
"""

import subprocess
import os
import argparse
import glob
import shutil
import datetime

# ── Excel bağımlılıkları ─────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Sabitler ──────────────────────────────────────────────────────────────────
IGNORE_LIST = ["fixed_ssdna.pdb", "initial_linear_ssdna.pdb"]
DOCKING_DIR = "Docking_result"
RAW_DIR = os.path.join(DOCKING_DIR, "raw_data")
REPORT_DIR = DOCKING_DIR  # Excel tablosu doğrudan bu klasöre konacak

# ── Renk Paleti ───────────────────────────────────────────────────────────────
CLR_HEADER_BG   = "1B2A4A"   # koyu lacivert
CLR_HEADER_FG   = "FFFFFF"
CLR_ACCENT      = "3B82F6"   # mavi aksan
CLR_GOOD        = "10B981"   # yeşil  (iyi skor)
CLR_WARN        = "F59E0B"   # sarı   (orta skor)
CLR_BAD         = "EF4444"   # kırmızı (kötü skor)
CLR_LIGHT_ROW   = "F1F5F9"
CLR_WHITE       = "FFFFFF"
CLR_TITLE_BG    = "0F172A"
CLR_SUBTITLE    = "64748B"
CLR_BORDER      = "CBD5E1"


# ═══════════════════════════════════════════════════════════════════════════════
#  1) APTAMER / RECEPTOR OTOMATİK TESPİT
# ═══════════════════════════════════════════════════════════════════════════════

def detect_ligand_and_receptors(manual_ligand=None, manual_receptors=None):
    """Klasördeki .pdb dosyalarını tarar.
    Adında 'APTAMER' (büyük harf) geçen dosyayı ligand olarak seçer.
    Geri kalanlar receptor olur."""

    all_pdbs = glob.glob("*.pdb")

    # ── Ligand belirleme ──
    if manual_ligand:
        ligand = manual_ligand
    else:
        # Aptamer32'yi otomatik olarak seç
        if os.path.exists("Aptamer32.pdb"):
            ligand = "Aptamer32.pdb"
        else:
            aptamer_files = [f for f in all_pdbs if "APTAMER" in f.upper()]
            if len(aptamer_files) == 1:
                ligand = aptamer_files[0]
            elif len(aptamer_files) > 1:
                print("⚠  Birden fazla APTAMER dosyası bulundu:")
                for i, af in enumerate(aptamer_files, 1):
                    print(f"   {i}. {af}")
                ligand = aptamer_files[0]
                print(f"   → İlk dosya seçildi: {ligand}")
            else:
                print("❌ Klasörde adında 'APTAMER' geçen bir .pdb dosyası bulunamadı!")
                print("   Lütfen -l parametresiyle ligand dosyasını belirtin.")
                return None, []

    if not os.path.exists(ligand):
        print(f"❌ Ligand dosyası '{ligand}' bulunamadı!")
        return None, []

    # ── Receptor belirleme ──
    if manual_receptors:
        receptors = manual_receptors
    else:
        receptors = [
            f for f in all_pdbs
            if f != ligand and f not in IGNORE_LIST
        ]
        receptors.sort()

    return ligand, receptors


# ═══════════════════════════════════════════════════════════════════════════════
#  2) HDOCK ÇALIŞTIRMA
# ═══════════════════════════════════════════════════════════════════════════════

def run_hdock(receptor, ligand):
    """Tek bir receptor için HDOCK çalıştırır. Başarı durumunu döndürür."""
    out_file = f"{os.path.splitext(receptor)[0]}_result.out"
    print(f"  🔬 Docking: {receptor} ↔ {ligand} ...")

    cmd_wsl = ['wsl', './hdock', receptor, ligand, '-out', out_file]
    cmd_native = ['./hdock', receptor, ligand, '-out', out_file]

    # WSL ile dene
    success = False
    try:
        proc = subprocess.run(cmd_wsl, stdout=subprocess.DEVNULL,
                              stderr=subprocess.PIPE, text=True)
        if proc.returncode == 0:
            success = True
    except Exception:
        pass

    # Başarısızsa native dene
    if not success:
        try:
            proc = subprocess.run(cmd_native, stdout=subprocess.DEVNULL,
                                  stderr=subprocess.PIPE, text=True)
            if proc.returncode == 0:
                success = True
            else:
                err = proc.stderr.strip()
                if err:
                    print(f"     ⚠ hdock hatası: {err}")
        except Exception as e:
            print(f"     ⚠ hdock çalıştırılamadı: {e}")

    if success:
        print(f"     ✅ Başarılı → {out_file}")
    else:
        print(f"     ❌ Başarısız: {receptor}")

    return success, out_file


# ═══════════════════════════════════════════════════════════════════════════════
#  3) SONUÇ PARSE ETME
# ═══════════════════════════════════════════════════════════════════════════════

def parse_hdock_results(out_file, top_n=5):
    """HDOCK .out dosyasından docking sonuçlarını parse eder.
    Her satır bir dict: {rank, score, rmsd, ...}"""

    results = []

    # Dosya ana dizinde veya raw_data içinde olabilir
    if not os.path.exists(out_file):
        alt = os.path.join(RAW_DIR, os.path.basename(out_file))
        if os.path.exists(alt):
            out_file = alt
        else:
            return results

    with open(out_file, 'r', encoding='utf-8', errors='ignore') as f:
        lines = f.readlines()

    count = 0
    for line in lines:
        parts = line.strip().split()
        # HDOCK çıktısında 9 sütunlu veri satırları
        if len(parts) == 9:
            try:
                floats = [float(p) for p in parts]
                count += 1
                results.append({
                    "rank": count,
                    "rot1": floats[0], "rot2": floats[1], "rot3": floats[2],
                    "tr1": floats[3], "tr2": floats[4], "tr3": floats[5],
                    "score": floats[6],
                    "rmsd": floats[7],
                })
                if count >= top_n:
                    break
            except ValueError:
                continue
        # Alternatif format: 3+ sütun, ilki rank (integer string)
        elif len(parts) >= 3 and parts[0].isdigit() and count == 0:
            try:
                count += 1
                results.append({
                    "rank": int(parts[0]),
                    "score": float(parts[1]),
                    "rmsd": float(parts[2]) if len(parts) > 2 else None,
                })
                if count >= top_n:
                    break
            except ValueError:
                continue

    return results


# ═══════════════════════════════════════════════════════════════════════════════
#  4) PROFESYONEL EXCEL RAPORU
# ═══════════════════════════════════════════════════════════════════════════════

def _apply_score_fill(cell, score):
    """Skora göre hücreyi renklendirir (düşük = iyi = yeşil)."""
    if score is None:
        return
    if score < -200:
        cell.fill = PatternFill("solid", fgColor=CLR_GOOD)
        cell.font = Font(color="FFFFFF", bold=True, size=11)
    elif score < -100:
        cell.fill = PatternFill("solid", fgColor=CLR_WARN)
        cell.font = Font(color="000000", bold=True, size=11)
    else:
        cell.fill = PatternFill("solid", fgColor=CLR_BAD)
        cell.font = Font(color="FFFFFF", bold=True, size=11)


def generate_excel_report(all_results, ligand_name, top_n=5):
    """Tüm docking sonuçlarını tek bir profesyonel Excel sayfasına yazar.
    
    all_results: dict  →  { "receptor_name": [list of result dicts], ... }
    """

    if not HAS_OPENPYXL:
        print("⚠  openpyxl yüklü değil. Excel raporu oluşturulamıyor.")
        print("   Yüklemek için:  pip install openpyxl")
        return None

    os.makedirs(REPORT_DIR, exist_ok=True)
    excel_path = os.path.join(REPORT_DIR, "Docking_Report.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Docking Results"

    # Stiller
    thin_border = Border(
        left=Side(style='thin', color=CLR_BORDER),
        right=Side(style='thin', color=CLR_BORDER),
        top=Side(style='thin', color=CLR_BORDER),
        bottom=Side(style='thin', color=CLR_BORDER),
    )
    header_font = Font(name="Segoe UI", bold=True, color=CLR_HEADER_FG, size=11)
    header_fill = PatternFill("solid", fgColor=CLR_HEADER_BG)
    title_font = Font(name="Segoe UI", bold=True, color=CLR_HEADER_FG, size=16)
    title_fill = PatternFill("solid", fgColor=CLR_TITLE_BG)
    subtitle_font = Font(name="Segoe UI", color=CLR_SUBTITLE, size=10, italic=True)
    data_font = Font(name="Segoe UI", size=11)
    accent_font = Font(name="Segoe UI", bold=True, color=CLR_ACCENT, size=12)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    receptors = list(all_results.keys())
    n_receptors = len(receptors)

    # ── Sütun yapısı ──────────────────────────────────────
    # A: Receptor Adı
    # B: Metric (Score / RMSD)
    # C..C+top_n-1: Rank 1, 2, 3 ...
    # Son sütun: Best Score / Min RMSD
    total_cols = 2 + top_n + 1  # A + B + ranks + summary
    last_col_letter = get_column_letter(total_cols)

    # ═══ BAŞLIK BÖLÜMÜ ═══════════════════════════════════
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    title_cell = ws.cell(row=row, column=1, value="🧬  HDOCK DOCKING REPORT")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 40

    row = 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    sub_text = (
        f"Ligand: {ligand_name}  |  "
        f"Receptors: {n_receptors}  |  "
        f"Top-{top_n} Results  |  "
        f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    sub_cell = ws.cell(row=row, column=1, value=sub_text)
    sub_cell.font = subtitle_font
    sub_cell.fill = PatternFill("solid", fgColor="F8FAFC")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22

    # Boş satır
    row = 3
    ws.row_dimensions[row].height = 8

    # ═══ RENK AÇIKLAMASI (LEGEND) ═══════════════════════
    row = 4
    ws.cell(row=row, column=1, value="Skor Renk Kılavuzu:").font = Font(
        name="Segoe UI", bold=True, size=10, color="334155"
    )
    
    legend_items = [
        (2, "🟢  Güçlü Bağlanma (< -200)", CLR_GOOD, "FFFFFF"),
        (3, "🟡  Orta Bağlanma (-200 ~ -100)", CLR_WARN, "000000"),
        (4, "🔴  Zayıf Bağlanma (> -100)", CLR_BAD, "FFFFFF"),
    ]
    for col_offset, text, bg, fg in legend_items:
        c = ws.cell(row=row, column=col_offset, value=text)
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name="Segoe UI", size=9, color=fg, bold=True)
        c.alignment = left

    row = 5
    ws.row_dimensions[row].height = 8

    # ═══ TABLO BAŞLIKLARI ═════════════════════════════════
    row = 6
    headers = ["RECEPTOR", "METRIC"]
    for i in range(1, top_n + 1):
        headers.append(f"RANK {i}")
    headers.append("BEST")

    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin_border
    ws.row_dimensions[row].height = 28

    # ═══ VERİ SATIRLARI ══════════════════════════════════
    row = 7
    for r_idx, receptor_name in enumerate(receptors):
        results = all_results[receptor_name]
        display_name = os.path.splitext(receptor_name)[0].replace("_", " ").title()

        is_even = r_idx % 2 == 0
        row_fill = PatternFill("solid", fgColor=CLR_LIGHT_ROW if is_even else CLR_WHITE)

        # ── Satır 1: Docking Score ──
        # Receptor adı (2 satırı kapsar)
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=1)
        rc = ws.cell(row=row, column=1, value=display_name)
        rc.font = accent_font
        rc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        rc.fill = row_fill
        rc.border = thin_border

        # Metric label
        mc = ws.cell(row=row, column=2, value="Docking Score")
        mc.font = Font(name="Segoe UI", bold=True, size=10, color="334155")
        mc.alignment = center
        mc.fill = row_fill
        mc.border = thin_border

        scores = []
        for rank_idx in range(top_n):
            col = 3 + rank_idx
            if rank_idx < len(results) and results[rank_idx].get("score") is not None:
                val = results[rank_idx]["score"]
                scores.append(val)
                c = ws.cell(row=row, column=col, value=round(val, 2))
                c.number_format = '0.00'
                _apply_score_fill(c, val)
            else:
                c = ws.cell(row=row, column=col, value="—")
                c.font = Font(name="Segoe UI", size=11, color="94A3B8")
            c.alignment = center
            c.border = thin_border
            if not c.fill or c.fill.fgColor is None or c.fill.fgColor.rgb == '00000000':
                c.fill = row_fill

        # Best score
        best_col = 3 + top_n
        if scores:
            best_score = min(scores)
            bc = ws.cell(row=row, column=best_col, value=round(best_score, 2))
            bc.number_format = '0.00'
            _apply_score_fill(bc, best_score)
        else:
            bc = ws.cell(row=row, column=best_col, value="—")
            bc.font = data_font
        bc.alignment = center
        bc.border = thin_border

        # ── Satır 2: RMSD ──
        row += 1
        mc2 = ws.cell(row=row, column=2, value="RMSD (Å)")
        mc2.font = Font(name="Segoe UI", bold=True, size=10, color="334155")
        mc2.alignment = center
        mc2.fill = row_fill
        mc2.border = thin_border

        rmsds = []
        for rank_idx in range(top_n):
            col = 3 + rank_idx
            if rank_idx < len(results) and results[rank_idx].get("rmsd") is not None:
                val = results[rank_idx]["rmsd"]
                rmsds.append(val)
                c = ws.cell(row=row, column=col, value=round(val, 2))
                c.number_format = '0.00'
                c.font = data_font
            else:
                c = ws.cell(row=row, column=col, value="—")
                c.font = Font(name="Segoe UI", size=11, color="94A3B8")
            c.alignment = center
            c.fill = row_fill
            c.border = thin_border

        # Min RMSD
        if rmsds:
            min_rmsd = min(rmsds)
            bc2 = ws.cell(row=row, column=best_col, value=round(min_rmsd, 2))
            bc2.number_format = '0.00'
            bc2.font = Font(name="Segoe UI", bold=True, size=11, color=CLR_ACCENT)
        else:
            bc2 = ws.cell(row=row, column=best_col, value="—")
            bc2.font = data_font
        bc2.alignment = center
        bc2.fill = row_fill
        bc2.border = thin_border

        row += 1

    # ═══ ÖZET BÖLÜMÜ ═════════════════════════════════════
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
    summary_title = ws.cell(row=row, column=1, value="📊  ÖZET KARŞILAŞTIRMA")
    summary_title.font = Font(name="Segoe UI", bold=True, size=13, color=CLR_HEADER_FG)
    summary_title.fill = PatternFill("solid", fgColor=CLR_ACCENT)
    summary_title.alignment = center
    ws.row_dimensions[row].height = 32

    row += 1
    sum_headers = ["RECEPTOR", "EN İYİ SKOR", "EN İYİ RMSD (Å)", "SONUÇ SAYISI", "DEĞERLENDİRME"]
    for col_idx, h in enumerate(sum_headers, 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin_border

    row += 1
    # İlk best_score'u bul (en iyi bağlanma) sıralama için
    summary_data = []
    for receptor_name in receptors:
        results = all_results[receptor_name]
        scores = [r["score"] for r in results if r.get("score") is not None]
        rmsds = [r["rmsd"] for r in results if r.get("rmsd") is not None]
        best_s = min(scores) if scores else None
        best_r = min(rmsds) if rmsds else None
        count = len(results)

        if best_s is not None and best_s < -200:
            verdict = "🟢 Güçlü Aday"
        elif best_s is not None and best_s < -100:
            verdict = "🟡 Potansiyel Aday"
        elif best_s is not None:
            verdict = "🔴 Zayıf Etkileşim"
        else:
            verdict = "⚪ Veri Yok"

        summary_data.append((receptor_name, best_s, best_r, count, verdict))

    # En iyi skora göre sırala
    summary_data.sort(key=lambda x: x[1] if x[1] is not None else 0)

    for s_idx, (rname, best_s, best_r, cnt, verdict) in enumerate(summary_data):
        display = os.path.splitext(rname)[0].replace("_", " ").title()
        is_even = s_idx % 2 == 0
        row_fill = PatternFill("solid", fgColor=CLR_LIGHT_ROW if is_even else CLR_WHITE)

        c1 = ws.cell(row=row, column=1, value=display)
        c1.font = Font(name="Segoe UI", bold=True, size=11)
        c1.alignment = left
        c1.fill = row_fill
        c1.border = thin_border

        c2 = ws.cell(row=row, column=2, value=round(best_s, 2) if best_s else "—")
        if best_s:
            c2.number_format = '0.00'
            _apply_score_fill(c2, best_s)
        c2.alignment = center
        c2.border = thin_border

        c3 = ws.cell(row=row, column=3, value=round(best_r, 2) if best_r else "—")
        if best_r:
            c3.number_format = '0.00'
        c3.font = data_font
        c3.alignment = center
        c3.fill = row_fill
        c3.border = thin_border

        c4 = ws.cell(row=row, column=4, value=cnt)
        c4.font = data_font
        c4.alignment = center
        c4.fill = row_fill
        c4.border = thin_border

        c5 = ws.cell(row=row, column=5, value=verdict)
        c5.font = Font(name="Segoe UI", bold=True, size=11)
        c5.alignment = center
        c5.fill = row_fill
        c5.border = thin_border

        row += 1

    # ═══ SÜTUN GENİŞLİKLERİ ══════════════════════════════
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 16
    for i in range(3, 3 + top_n):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.column_dimensions[get_column_letter(3 + top_n)].width = 14

    # Sayfa ayarları
    ws.sheet_properties.tabColor = CLR_ACCENT
    ws.freeze_panes = "A7"

    # Kaydet
    wb.save(excel_path)
    return excel_path


# ═══════════════════════════════════════════════════════════════════════════════
#  5) HAM DOSYALARI TAŞIMA
# ═══════════════════════════════════════════════════════════════════════════════

def move_raw_files(out_files):
    """Üretilen .out dosyalarını raw_data klasörüne taşır."""
    os.makedirs(RAW_DIR, exist_ok=True)
    for f in out_files:
        if os.path.exists(f):
            dest = os.path.join(RAW_DIR, os.path.basename(f))
            shutil.move(f, dest)


# ═══════════════════════════════════════════════════════════════════════════════
#  6) ANA FONKSİYON
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="HDOCK Docking & Report Pipeline",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Örnekler:
  python run_hdock.py                         Otomatik: APTAMER → ligand, geri kalan → receptor
  python run_hdock.py -l OzelLigand.pdb       Belirli bir ligand kullan
  python run_hdock.py -r hedef1.pdb hedef2.pdb  Sadece belirli receptorlarla dock et
  python run_hdock.py --top 10                Her receptor için en iyi 10 sonucu al
        """
    )
    parser.add_argument("-l", "--ligand", help="Ligand PDB dosyası (varsayılan: APTAMER içeren dosya)")
    parser.add_argument("-r", "--receptors", nargs='*', help="Receptor PDB dosyaları")
    parser.add_argument("--top", type=int, default=5, help="Her receptor için kaç sonuç alınsın (varsayılan: 5)")
    parser.add_argument("--skip-dock", action="store_true",
                        help="Docking'i atla, sadece mevcut .out dosyalarından rapor üret")
    args = parser.parse_args()

    print("=" * 60)
    print("  🧬  HDOCK DOCKING & REPORT PIPELINE")
    print("=" * 60)

    # ── Dosya tespiti ──
    ligand, receptors = detect_ligand_and_receptors(args.ligand, args.receptors)
    if not ligand or not receptors:
        return

    print(f"\n  📌  Ligand   : {ligand}")
    print(f"  🎯  Receptors: {len(receptors)} dosya")
    for r in receptors:
        print(f"       • {r}")
    print()

    out_files = []

    # ── Docking ──
    if not args.skip_dock:
        print("─" * 60)
        print("  ADIM 1: HDOCK Docking")
        print("─" * 60)
        for receptor in receptors:
            success, out_file = run_hdock(receptor, ligand)
            out_files.append(out_file)
        print()

    # ── Parse ──
    print("─" * 60)
    print("  ADIM 2: Sonuçları Parse Etme")
    print("─" * 60)

    all_results = {}
    for receptor in receptors:
        out_file = f"{os.path.splitext(receptor)[0]}_result.out"
        results = parse_hdock_results(out_file, top_n=args.top)
        all_results[receptor] = results

        if results:
            best = min(r["score"] for r in results if r.get("score") is not None)
            print(f"  ✅ {receptor}: {len(results)} sonuç bulundu (en iyi skor: {best:.2f})")
        else:
            print(f"  ⚠  {receptor}: Sonuç bulunamadı")
    print()

    # ── Rapor ──
    print("─" * 60)
    print("  ADIM 3: Excel Raporu Oluşturma")
    print("─" * 60)

    excel_path = generate_excel_report(all_results, ligand, top_n=args.top)
    if excel_path:
        print(f"  📊  Rapor oluşturuldu: {excel_path}")
    print()

    # ── Taşıma ──
    if out_files:
        move_raw_files(out_files)
        print(f"  📁  Ham veriler taşındı: {RAW_DIR}")

    print("=" * 60)
    print("  ✅  PIPELINE TAMAMLANDI")
    print("=" * 60)


if __name__ == "__main__":
    main()
